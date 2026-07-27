from typing import Any, List, Dict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formula.tokenizer import Tokenizer
import io
from ..parser import BaseParser, PARSER_REGISTRY

class ExcelParser(BaseParser):
    def __init__(self):
        self.semantic_map = {
            "Revenue": ["rev", "revenue", "sales amount", "net sales", "income", "turnover"],
            "Customer ID": ["customer_id", "client_number", "account_code", "user id"],
            "Date": ["date", "timestamp", "created_at", "updated_at"],
            "Product": ["product", "item", "sku"],
            "Cost": ["cost", "cogs", "cost of goods sold"],
            "Quantity": ["quantity", "qty", "units"],
            "Price": ["price", "unit price"],
        }

    def _semantic_analysis(self, column_name: str) -> str | None:
        for concept, aliases in self.semantic_map.items():
            if column_name.lower() in aliases:
                return concept
        return None

    def _extract_formulas(self, wb):
        formulas = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f':
                        tok = Tokenizer(cell.value)
                        formulas.append({
                            "sheet": sheet.title,
                            "cell": cell.coordinate,
                            "formula": cell.value,
                            "tokens": [t.value for t in tok.items],
                        })
        return formulas

    def parse(self, file_content: bytes) -> Dict[str, Any]:
        # Use a file-like object for pandas
        file_like_object = io.BytesIO(file_content)
        xls = pd.ExcelFile(file_like_object, engine='openpyxl')
        sheets = {}
        for sheet_name in xls.sheet_names:
            sheets[sheet_name] = xls.parse(sheet_name)
        
        # Use a file-like object for openpyxl
        file_like_object.seek(0)
        wb = load_workbook(file_like_object, data_only=False)
        metadata = {
            "sheet_names": wb.sheetnames,
            "hidden_sheets": [sheet.title for sheet in wb.worksheets if sheet.sheet_state == 'hidden'],
            "named_ranges": [nr.name for nr in wb.defined_names.definedName],
            "tables": [],
            "formulas": self._extract_formulas(wb),
        }
        for ws in wb.worksheets:
            for table in ws.tables.values():
                metadata["tables"].append({"name": table.name, "range": table.ref})
        
        return {"sheets": sheets, "metadata": metadata}

    def extract_content(self, parsed_data: Dict[str, Any]) -> str:
        content = []
        for sheet_name, df in parsed_data["sheets"].items():
            content.append(f"Sheet: {sheet_name}\n")
            content.append(df.to_string())
        return "\n".join(content)

    def extract_metadata(self, parsed_data: Dict[str, Any]) -> Dict[str, Any]:
        metadata = parsed_data.get("metadata", {})
        sheets_analysis = {}
        for sheet_name, df in parsed_data["sheets"].items():
            sheets_analysis[sheet_name] = {
                "columns": list(df.columns),
                "dtypes": {col: str(dtype) for col, dtype in df.dtypes.items()},
                "rows": len(df),
                "duplicate_rows": int(df.duplicated().sum()),
                "missing_values": int(df.isnull().sum().sum()),
            }
        metadata["sheets_analysis"] = sheets_analysis
        return metadata

    def extract_entities(self, parsed_data: Dict[str, Any]) -> List[Dict[str, Any]]:
        entities = []
        for sheet_name, df in parsed_data["sheets"].items():
            for col in df.columns:
                business_concept = self._semantic_analysis(col)
                if business_concept:
                    entities.append({
                        "entity": business_concept,
                        "value": col,
                        "context": f"Found in sheet '{sheet_name}'"
                    })
        return entities

    def extract_relationships(self, parsed_data: Dict[str, Any]) -> List[Dict[str, Any]]:
        return []

    def validate(self, parsed_data: Dict[str, Any]) -> bool:
        return "sheets" in parsed_data

def initialize():
    PARSER_REGISTRY.register_parser(".xlsx", ExcelParser())
    PARSER_REGISTRY.register_parser(".xls", ExcelParser())

initialize()
